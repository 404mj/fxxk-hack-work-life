import openpyxl

# https://blog.csdn.net/sinat_28576553/article/details/81275650  || https://zhuanlan.zhihu.com/p/61059318
# https://www.linuxzen.com/python-you-ya-de-cao-zuo-zi-dian.html

# 构造人员信息属性
mem_attr={}
lev2_attr=dict()
lev2_attr["李世冲"]=["维护室","大数据中心"]
lev2_attr["李辉"]=["服务室","业务室"]
lev2_attr["程进"]=["管信室","运营室"]
meminfoWb=openpyxl.load_workbook("xxjsb_list_nonewbie_leader.xlsx")
sht_memlist=meminfoWb.worksheets[1]

for i in range(2,76):
    levstr=sht_memlist.cell(i,4).value
    memname=sht_memlist.cell(i,3).value
    keshi=sht_memlist.cell(i,2).value
    if(levstr.strip()=="员工"):
	    level=4
    elif(levstr.strip()=="三级"):
	    level=3
    elif(levstr.strip()=="副总"):
	    level=2
    else:
	    level=1
	# 姓名：[科室，级别]
    mem_attr[memname]=[keshi,level]


# 原始高职级排名
senOrder={}
oriOrderWb=openpyxl.load_workbook("3seasonOriOrder.xlsx")
sht_order=oriOrderWb.worksheets[0]
for i in range(1,43):
	senOrder[sht_order.cell(i,1).value]=sht_order.cell(i,2).value

#原始低职级排名
lowOrder={}
lowWb=openpyxl.load_workbook("3seasonLowOriorder.xlsx")
sht_low=lowWb.worksheets[0]
for i in range(1,23):
	lowOrder[sht_low.cell(i,1).value]=sht_low.cell(i,2).value

# 处理高职级评价
wb=openpyxl.load_workbook("20202finalSenior.xlsx")
sht=wb.worksheets[0]
for i in range(2,75):
	judger=sht.cell(2,i).value
	judgerLevel=mem_attr[judger][1]
	judgerKeshi=mem_attr[judger][0]
	for j in range(3,45):
		rate=1
		order_name=sht.cell(j,1).value
		if judgerLevel==4:
			# 同为员工相同科室
			if judgerKeshi==mem_attr[order_name][0]:
				rate=2
		elif judgerLevel==1:
		    rate=10
		elif judgerLevel==3:
			# 王新印分管两个科室
			if mem_attr[order_name][0] in judgerKeshi:
				rate=4
			else:
				rate=2
		else:
			# 评价人为副总看被评价人是否在其分管范围下
			if mem_attr[order_name][0] in lev2_attr[judger]:
				rate=8
			else:
				rate=4
		# 李玉亮 朱祥磊
		if judgerKeshi=="规划组":
			rate = 2
		# 评价自己权重为1
		if judger==order_name:
			rate = 1
		# 写入权重
		if sht.cell(j,i).value=="-":
			sht.cell(j,i).value=0
		else:
			sht.cell(j,i).value=rate*sht.cell(j,i).value
wb.save("20202finalSenior.xlsx")

# 处理低评高减半
wb=openpyxl.load_workbook("20202finalSenior.xlsx")
sht=wb.worksheets[0]
for i in range(2,75):
	judger=sht.cell(2,i).value
	for j in range(3,45):
		# 写如减半操作
		if judger in lowOrder:
			sht.cell(j,i).value=sht.cell(j,i).value*0.5
wb.save("20202finalSenior.xlsx")



# 处理低职级评价
wbgood=openpyxl.load_workbook("20202finalGood.xlsx")
shtgood=wbgood.worksheets[0]
for i in range(2,75):
	judger=shtgood.cell(2,i).value
	judgerLevel=mem_attr[judger][1]
	judgerKeshi=mem_attr[judger][0]
	for j in range(3,25):
		rate=1
		order_name=shtgood.cell(j,1).value
		if judgerLevel==4:
			# 同为员工相同科室
			if judgerKeshi==mem_attr[order_name][0]:
				rate=2
		elif judgerLevel==1:
		    rate=10
		elif judgerLevel==3:
			if mem_attr[order_name][0] in judgerKeshi:
				rate=4
			else:
				rate=2
		else:
			# 评价人为副总看被评价人是否在其分管范围下
			if mem_attr[order_name][0] in lev2_attr[judger]:
				rate=8
			else:
				rate=4
		# 李玉亮 朱祥磊
		if judgerKeshi=="规划组":
			rate = 2
		# 评价自己权重为1
		if judger==order_name:
			rate = 1
		# 写入权重
		if shtgood.cell(j,i).value=="-":
			shtgood.cell(j,i).value=0
		else:
			shtgood.cell(j,i).value=rate*shtgood.cell(j,i).value
wbgood.save("20202finalGood.xlsx")


# 处理高评低减半
wbgood=openpyxl.load_workbook("20202finalGood.xlsx")
shtgood=wbgood.worksheets[0]
for i in range(2,75):
	judger=shtgood.cell(2,i).value
	for j in range(3,25):
		# 写如减半操作
		if judger in senOrder:
			shtgood.cell(j,i).value=shtgood.cell(j,i).value*0.5
wbgood.save("20202finalGood.xlsx")

# 由于权重增加为2，4，8，10，故将排名反转，计算高分
# 注：互评应该将自己放到原始位置不变！否则别人多了22分。
for i in range(2,75):
	name=sht.cell(2,i).value
	for j in range(3,45):
		cellv=sht.cell(j,i).value

		if name not in senOrder:
			sht.cell(j,i).value=43-cellv
		else:
			if cellv=="-":
				sht.cell(j,i).value=43-senOrder[name]
			elif cellv < senOrder[name]:
				sht.cell(j,i).value=43-cellv
			elif cellv >= senOrder[name]:
				sht.cell(j,i).value=43-cellv-1
wb.save("20202finalSenior.xlsx")

# 按照以上逻辑处理低职级排名
for i in range(2,75):
	name=shtgood.cell(2,i).value
	for j in range(3,25):
		cellv=shtgood.cell(j,i).value
		if name not in lowOrder:
			shtgood.cell(j,i).value=23-cellv
		else:
			if cellv=="-":
				shtgood.cell(j,i).value=23-lowOrder[name]
			elif cellv < lowOrder[name]:
				shtgood.cell(j,i).value=23-cellv
			elif cellv >= lowOrder[name]:
				shtgood.cell(j,i).value=23-cellv-1
wbgood.save("20202finalGood.xlsx")



wb=openpyxl.load_workbook("20202finalSenior.xlsx")
sht=wb.worksheets[0]
# 每个科室内部排名-高职级
wbks=openpyxl.Workbook()
wsks = wbks.active
nowKeshi="大数据中心"
nowfile="20202大数据.xlsx"
m,n=2,2
for i in range(2,75):
	judger=sht.cell(2,i).value
	judgerKeshi=mem_attr[judger][0]
	if judgerKeshi==nowKeshi:
		wsks.cell(1,n).value=judger
		m=2
		for j in range(3,45):
			order_name=sht.cell(j,1).value
			if judgerKeshi==mem_attr[order_name][0]:
				wsks.cell(m,1).value=order_name
				wsks.cell(m,n).value=sht.cell(j,i).value
				m+=1
		n+=1
wbks.save(nowfile)
# 处理低职级科室内部
wbgood=openpyxl.load_workbook("20202finalGood.xlsx")
shtgood=wbgood.worksheets[0]
wbks=openpyxl.load_workbook(nowfile)
wsks=wbks.worksheets[1]
m,n=2,2
for i in range(2,75):
	judger=shtgood.cell(2,i).value
	judgerKeshi=mem_attr[judger][0]
	if judgerKeshi==nowKeshi:
		wsks.cell(1,n).value=judger
		m=2
		for j in range(3,25):
			order_name=shtgood.cell(j,1).value
			if judgerKeshi==mem_attr[order_name][0]:
				wsks.cell(m,1).value=order_name
				wsks.cell(m,n).value=shtgood.cell(j,i).value
				m+=1
		n+=1
wbks.save(nowfile)


wbnew.save("20202大数据.xlsx")



