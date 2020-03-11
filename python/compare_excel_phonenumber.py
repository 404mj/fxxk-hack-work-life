from openpyxl import load_workbook
wberr=load_workbook('./20200311省公司安卓十且低版本.xlsx')
wbtem=load_workbook('./省公司上传模板20200311.xlsx')
errsht=wberr.active
temsht=wbtem.active
errlist=[]
temlist=[]
for x in range(2,96):
    errlist.append(errsht.cell(x,4).value)
for y inrange(4,728):
    temlist.append(str(temsht.cell(y,5).value))
errset=set(errlist)
temset=set(temlist)
preset=errset&temset
len(preset)