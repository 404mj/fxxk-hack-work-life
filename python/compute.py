import openpyxl

# https://blog.csdn.net/sinat_28576553/article/details/81275650  || https://zhuanlan.zhihu.com/p/61059318
# https://www.linuxzen.com/python-you-ya-de-cao-zuo-zi-dian.html

wb=openpyxl.load_workbook("20202pre.xlsx")
sht0=wb.worksheets[0]
sht1=wb.worksheets[1]
sen_ori={}
for i in range(2,44):
	sen_ori[sht1.cell(i,1).value]=i


mem_attr={}
# lev3_attr={"宋海峰":"业务室","徐振龙":"服务室","刘洪波":"管信室","王新印":"运营室","田力":"维护室","王新印":"大数据中心"}
lev2_attr["李世冲"]=["维护室","大数据中心"]
lev2_attr["李辉"]=["服务室","业务室"]
lev2_attr["程进"]=["管信室","运营室"]
wb2=openpyxl.load_workbook("xxjsb_list_nonewbie_leader.xlsx")
sht_memlist=wb2.worksheets[1]

for i in range(2,75):
    levstr=sht_memlist.cell(i,4).value
    if(levstr.strip()=="员工"):
	    level=4
    elif(levstr.strip()=="三级"):
	    level=3
    elif(levstr.strip()=="副总"):
	    level=2
    else:
	    level=1
    mem_attr[sht_memlist.cell(i,3).value]=[sht_memlist.cell(i,2).value,level]


for i in range(2,64):
	name=sht0.cell(2,i).value
	level=mem_attr[name][1]
	keshi=mem_attr[name][0]
	for j in range(3,45):
		rate=1
		order_name=sht0.cell(j,i).value
		if level==4:
			# 同为员工相同科室
			if keshi==mem_attr[order_name][0]:
				rate=0.8
		elif level==1:
		    rate=0.5
		elif level==3:
			if keshi == mem_attr[order_name][0]:
				rate=0.5
			else:
				rate=0.7
		else:
			# 评价人为副总看被评价人是否在其分管范围下
			if mem_attr[order_name][0] in lev2_attr[name]:
				rate=0.5
			else:
				rate=0.6

		# 写到另一个sheet中	
		sht1.cell(sen_ori[order_name],2).value=rate*(j-2)
	sht1.cell(1,i).value=name
wb.save("20202pre.xlsx")		





