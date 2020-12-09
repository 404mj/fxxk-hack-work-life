import os
from itertools import groupby
# 处理六楼测试照片
import openpyxl

inxls=openpyxl.load_workbook("6Lceshi.xlsx")
insht=inxls.worksheets[0]
start_row=6

# https://juejin.im/post/5c57afb1f265da2dda6924a1
# 处理用户名-分割数字和其他。
nowdir = os.walk("./")
for path,dir_list,file_list in nowdir:
    for file_name in file_list:
        if not file_name.endswith(".xlsx"):
            ss = [''.join(list(g)) for k, g in groupby(file_name, key=lambda x: x.isdigit())]
            if ss[2].startswith("眼镜"):
                picext="."+ss[2].split(".")[1]
            else:
                picext=ss[2]
            insht.cell(start_row,1).value=ss[0]
            insht.cell(start_row,5).value=ss[1]
            os.rename(file_name,ss[1]+picext)
            start_row+=1
# 必须空一行
inxls.save("6Lpics.xlsx")


# 处理图片大小
from glob import glob
from PIL import Image

threshold=300*1024
for _,_,imgfile_list in nowdir:
    for imgname in imgfile_list:
        if os.path.getsize(imgname) >= threshold:
            img=Image.open(imgname)
            img=img.resize((585,760),Image.ANTIALIAS)
            img.save(imgname,quality=100)

files=os.listdir(".")
for file in files:
    portion=os.path.splitext(file)
    if portion[1]==".JPG":
        os.rename(file,portion[0]+".jpg")

#-------20201207 updated----------
import os
from PIL import Image

for _,_,imgfile_list in os.walk('./'):
    for imgname in imgfile_list:
        if os.path.getsize(imgname) >= 300000:
            img = Image.open(imgname)
            img.save("compressed_"+imgname,
             optimize=True,
             quality=30)

#--------20201207 updated 2---------
pip3 install pillow optimize-images

#----------处理文件名并写入excel------------
import os
import shutil
from itertools import groupby
import openpyxl

i=0
j=0
dir=0
xls='faceup'
for path, pathlist, filelist in os.walk('./'):
    for img in filelist:
        phone=img.strip().split('-')[0]
        name=img.strip.split('-')[1]
        newname=phone+'.jpg'
        os.rename(img,newname)
        i+=1
        if i%200 > 0:
            xlsname=xls+str(j)+'.xlsx'
            inxls=openpyxl.load_workbook("../faceup.xlsx")
            insht=inxls.worksheets[0]
            start_row=6 
            insht.cell(start_row,1).value=name 
            insht.cell(start_row,5).value=phone 
            newdir=os.path.join(os.getcwd(),str(dir))
            os.makedirs(newdir)
            shutil.move(os.getcwd()+'/'+newname,newdir+'/'+newname)
        else if :
            pass
# 无法实现自动命名且移动且写excel的逻辑。


#--------------重新实现--------------
import os
import shutil
from itertools import groupby
import openpyxl

#把文件名都读到二维list
imgs=[]
mid=[]
i=0
for path, pathlist, filelist in os.walk('./'):
    for img in filelist:
        mid.append(img)
        i+=1
        if i%2 == 0:
            imgs.append(mid)
            mid=[]
    imgs.append(mid)

for i in range(0,len(imgs)):
    xlsname='faceup'+str(i)+'.xlsx'
    subimg=imgs[i]
    wb=openpyxl.Workbook()
    ws=wb.active
    startrow=1
    newdir=os.path.join(os.getcwd(),str(i))
    os.makedirs(newdir)
    for j in range(0,len(subimg)):
        img=subimg[j]
        phone=img.strip().split('-')[0]
        name=img.strip().split('-')[1].split('.')[0]
        newname=phone+'.jpg'
        os.rename(img,newname)
        ws.cell(startrow,1).value=name
        ws.cell(startrow,5).value=phone 
        startrow+=1
        shutil.move(os.getcwd()+'/'+newname,newdir+'/'+newname)
    wb.save(xlsname)



