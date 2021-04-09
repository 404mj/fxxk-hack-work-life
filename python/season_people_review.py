import openpyxl

# https://blog.csdn.net/sinat_28576553/article/details/81275650  || https://zhuanlan.zhihu.com/p/61059318
# https://www.linuxzen.com/python-you-ya-de-cao-zuo-zi-dian.html

# 参与调查员工总数
NUM=75
# 高职级被评价人数
SENNUM=38
# 高职级的评价人数
SJ=56
# 低职级被评价人数
LOWNUM=25
# 低职级的评价人数
LJ=54

# 构造人员科室级别-管理对应关系
mem_attr={}
lev2_attr=dict()
lev2_attr["李世冲"]=["维护室","大数据"]
lev2_attr["李辉"]=["服务室","业务室"]
lev2_attr["程进"]=["管信室","运营室","安全室"]
meminfoWb=openpyxl.load_workbook("20211empolyes.xlsx")
memlist=meminfoWb.worksheets[1]

for i in range(2,NUM+2):
    levstr=memlist.cell(i,4).value
    memname=memlist.cell(i,3).value
    keshi=memlist.cell(i,2).value
    if(levstr.strip()=="员工"):
	    level=4
    elif(levstr.strip()=="三级"):
	    level=3
    elif(levstr.strip()=="副总"):
	    level=2
    else:
        level=1
	# 姓名：[科室，级别]
    mem_attr[memname]=[keshi,level]

# 构造高低职级人员原始排名顺序
senOrder={} #姓名:原始顺序
senOrderWb=openpyxl.load_workbook("SenOrder.xlsx")
sht_sen=senOrderWb.worksheets[0]
for i in range(1,SENNUM+1):
	senOrder[sht_sen.cell(i,1).value]=sht_sen.cell(i,2).value


lowOrder={}  #姓名:原始顺序
lowOrderWb=openpyxl.load_workbook("LowOrder.xlsx")
sht_low=lowOrderWb.worksheets[0]
for i in range(1,LOWNUM+1):
	lowOrder[sht_low.cell(i,1).value]=sht_low.cell(i,2).value



# ================>高职级流水处理
# 先反转:排名->得分
# 注：互评应该将自己放到原始位置不变！否则别人多了22分。
#定义好魔数: 被评价人数SENNUM=38,列range=SENNUM+3,REV=SENNUM+1,评价人数SJ=56,行range=SJ+1
EMPLOYE_RANGE_S=SENNUM+3
SREV=SENNUM+1
JUDGER_RANGE_S=SJ+1

wbs1=openpyxl.load_workbook("senior.xlsx")
shts1=wbs1.worksheets[0]
for i in range(2,EMPLOYE_RANGE_S):
    #评价人
    name=shts1.cell(2,i).value
    for j in range(3,JUDGER_RANGE_S):
        cellv=shts1.cell(j,i).value
        if name not in senOrder:
            shts1.cell(j,i).value=SREV-cellv
        else:
            if cellv=="-":
                shts1.cell(j,i).value=SREV-senOrder[name]
            elif cellv < senOrder[name]:
                shts1.cell(j,i).value=SREV-cellv
            elif cellv >= senOrder[name]:
                shts1.cell(j,i).value=SREV-cellv-1
wbs1.save("senior_rev.xlsx")

# 按照管理划分生熟 计算权重2，4，8，10
wbs2=openpyxl.load_workbook("senior_rev.xlsx")
shts2=wbs2.worksheets[0]
# 处理高职级
# 这里处理senior_rev.xlsx 新增了一个虚拟三级经理，所以要+1
for i in range(2,JUDGER_RANGE_S+1):
	judger=shts2.cell(2,i).value
	judgerLevel=mem_attr[judger][1]
	judgerKeshi=mem_attr[judger][0]
	for j in range(3,ORDER_RANGE_S):
		rate=1
		order_name=shts2.cell(j,1).value
		if judgerLevel==4:
			# 同为员工相同科室
			if judgerKeshi==mem_attr[order_name][0]:
				rate=2
		elif judgerLevel==1:
		    rate=10
		elif judgerLevel==3:
			# 三级经理评价
			if mem_attr[order_name][0] == judgerKeshi:
				rate=4
			else:
				rate=2
		else:
			# 评价人为副总看被评价人是否在其分管范围下
			if mem_attr[order_name][0] in lev2_attr[judger]:
				rate=8
			else:
				rate=4
		# 李玉亮 朱祥磊
		if judgerKeshi=="规划组":
			rate = 2
		# 评价自己权重为1
		if judger==order_name:
			rate = 1
		# 写入权重
		shts2.cell(j,i).value=rate*shts2.cell(j,i).value
wbs2.save("senior_mid.xlsx")

# 处理低评高减半
wbs3=openpyxl.load_workbook("senior_mid.xlsx")
shts3=wbs3.worksheets[0]
for i in range(2,JUDGER_RANGE_S):
	judger=shts3.cell(2,i).value
	if judger in lowOrder:
		for j in range(3,EMPLOYE_RANGE_S):
		# 写如减半操作
			shts3.cell(j,i).value=shts3.cell(j,i).value*0.5
wbs3.save("senior_final.xlsx")




# ================>低职级流水处理
wbl1=openpyxl.load_workbook("low.xlsx")
shtl1=wbl1.worksheets[0]
# 先反转:排名->得分
# 注：互评应该将自己放到原始位置不变！否则别人多了22分。
#定义好魔数: 被评价人数LOWNUM=38,列range=SC+3,LREV=LOWNUM+1,评价人数LJ=54,行range=LJ+1
EMPLOYE_RANGE_L=LOWNUM+3
LREV=LOWNUM+1
JUDGER_RANGE_L=LJ+1
for i in range(2,JUDGER_RANGE_L):
    name=shtl1.cell(2,i).value
    for j in range(3,EMPLOYE_RANGE_L):
        cellv=shtl1.cell(j,i).value
        if name not in lowOrder:
            shtl1.cell(j,i).value=LREV-cellv
        else:
            if cellv=="-":
                shtl1.cell(j,i).value=LREV-lowOrder[name]
            elif cellv < lowOrder[name]:
                shtl1.cell(j,i).value=LREV-cellv
            elif cellv >= lowOrder[name]:
                shtl1.cell(j,i).value=LREV-cellv-1
wbl1.save("low_rev.xlsx")

# 按照管理划分生熟 计算权重2，4，8，10
wbl2=openpyxl.load_workbook("low_rev.xlsx")
shtl2=wbl2.worksheets[0]
for i in range(2,JUDGER_RANGE_L+1):
	judger=shtl2.cell(2,i).value
	judgerLevel=mem_attr[judger][1]
	judgerKeshi=mem_attr[judger][0]
	for j in range(3,EMPLOYE_RANGE_L):
		rate=1
		order_name=shtl2.cell(j,1).value
		if judgerLevel==4:
			# 同为员工相同科室
			if judgerKeshi==mem_attr[order_name][0]:
				rate=2
		elif judgerLevel==1:
		    rate=10
		elif judgerLevel==3:
			if mem_attr[order_name][0] == judgerKeshi:
				rate=4
			else:
				rate=2
		else:
			# 评价人为副总看被评价人是否在其分管范围下
			if mem_attr[order_name][0] in lev2_attr[judger]:
				rate=8
			else:
				rate=4
		# 李玉亮 朱祥磊
		if judgerKeshi=="规划组":
			rate = 2
		# 评价自己权重为1
		if judger==order_name:
			rate = 1
		# 写入权重
		shtl2.cell(j,i).value=rate*shtl2.cell(j,i).value
wbl2.save("low_mid.xlsx")

# 处理高评低减半
wbl3=openpyxl.load_workbook("low_mid.xlsx")
shtl3=wbl3.worksheets[0]
for i in range(2,JUDGER_RANGE_L+1):
	judger=shtl3.cell(2,i).value
	for j in range(3,EMPLOYE_RANGE_L):
		# 写如减半操作
		if judger in senOrder:
			shtl3.cell(j,i).value=shtl3.cell(j,i).value*0.5
wbl3.save("low_final.xlsx")





# ================>三级经理流水处理
wb3ji=openpyxl.load_workbook("3jingli.xlsx")
sht3ji=wb3ji.worksheets[0]
for i in range(2,6):
	judger=sht3ji.cell(2,i).value
	for j in range(3,8):
		order_name=sht3ji.cell(j,1).value
		cellv=sht3ji.cell(j,i).value
		if mem_attr[judger][1] == 1:
			sht3ji.cell(j,i).value=(6-cellv)*10
		# 此处王新印受两个副总8倍计算
		else:
			if mem_attr[order_name][0] in lev2_attr[judger]:
				sht3ji.cell(j,i).value=(6-cellv)*8
			else:
				sht3ji.cell(j,i).value=(6-cellv)*4
wb3ji.save("3jinglifinal.xlsx")


# ===================================
wb=openpyxl.load_workbook("20202finalSenior.xlsx")
sht=wb.worksheets[0]
# 每个科室内部排名-高职级
wbks=openpyxl.Workbook()
wsks = wbks.active
nowKeshi="大数据中心"
nowfile="20202大数据.xlsx"
m,n=2,2
for i in range(2,75):
	judger=sht.cell(2,i).value
	judgerKeshi=mem_attr[judger][0]
	if judgerKeshi==nowKeshi:
		wsks.cell(1,n).value=judger
		m=2
		for j in range(3,45):
			order_name=sht.cell(j,1).value
			if judgerKeshi==mem_attr[order_name][0]:
				wsks.cell(m,1).value=order_name
				wsks.cell(m,n).value=sht.cell(j,i).value
				m+=1
		n+=1
wbks.save(nowfile)
# 处理低职级科室内部
wbgood=openpyxl.load_workbook("20202finalGood.xlsx")
shtgood=wbgood.worksheets[0]
wbks=openpyxl.load_workbook(nowfile)
wsks=wbks.worksheets[1]
m,n=2,2
for i in range(2,75):
	judger=shtgood.cell(2,i).value
	judgerKeshi=mem_attr[judger][0]
	if judgerKeshi==nowKeshi:
		wsks.cell(1,n).value=judger
		m=2
		for j in range(3,25):
			order_name=shtgood.cell(j,1).value
			if judgerKeshi==mem_attr[order_name][0]:
				wsks.cell(m,1).value=order_name
				wsks.cell(m,n).value=shtgood.cell(j,i).value
				m+=1
		n+=1
wbks.save(nowfile)
wbnew.save("20202大数据.xlsx")



